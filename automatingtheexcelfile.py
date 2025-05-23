import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
path=r'c:\Users\Vaishu\Downloads\train.csv'
df=pd.read_csv(path)
print(df)
#understanding the data
print(df.head(10))
print(df.tail(10))
print(df.shape)
print(df.info)#overview of colmns and null values
print(df.describe())
print(df.Pclass)
print(df[['Pclass','PassengerId']])
print(df.loc[5])
print(df.iloc[1:10])
#Data cleaning.
"Handling missing values"
print(df.isnull())
print(df.isnull().sum())
print(df.Age)
#since age is numerical we fill null values with the median
print(df['Age'].median())
df['Age']=df['Age'].fillna(value=28.0)
print(df['Age'])#inplace=true can be used
print(df.isnull().sum())#zero null values will be shown in the age colmn
#Embarked is a catogorical value so replacing the null values with repeated value
print(df.Embarked.mode()[0])#prints the most repeated value (s) herer
df['Embarked']=df['Embarked'].fillna(value="s")
print(df.isnull().sum())#zero null in embarked
#one more colmn has null values ie cabin
df=df.drop(columns=['Cabin'])
print(df)#it has more null values so simply dropped
"Handling Duplicates"
print(df.duplicated().sum())#no dupliactes are there so (0) simply
"Data type conversion"
print(df.dtypes)
df['Sex']=df['Sex'].map({'male':0,'female':1})#converting 
print(df.Sex)
df['Survived']=df['Survived'].astype('category')
df['Pclass']=df['Pclass'].astype('category')
df['Embarked']=df['Embarked'].astype('category')
print(df)
'Replacing the colmn names'
df=df.rename(columns={'Embarked':'Passenger_started'})
print(df)
'creating a new cleaned titanic excel file'
df.to_excel("cleaned_titanic.xlsx",index=False,engine='openpyxl')
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

# Load the workbook and data worksheet
wb = load_workbook("cleaned_titanic.xlsx")
ws = wb.active

# Format header row in data sheet
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")

# Adjust column widths for better readability
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = (max_length + 2) * 1.2

# Create a new worksheet for charts
chart_ws = wb.create_sheet(title="Charts")

# --- Survival Counts Data ---
chart_ws["A1"] = "Survived"
chart_ws["B1"] = "Count"

survived_counts = {0: 0, 1: 0}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    survived_counts[row[1]] += 1  # Survived column is index 1

chart_ws["A2"] = 0
chart_ws["B2"] = survived_counts[0]
chart_ws["A3"] = 1
chart_ws["B3"] = survived_counts[1]

# Survival Bar Chart
bar_chart = BarChart()
bar_chart.title = "Survival Counts"
bar_chart.x_axis.title = "Survived"
bar_chart.y_axis.title = "Count"
data = Reference(chart_ws, min_col=2, min_row=1, max_row=3)
cats = Reference(chart_ws, min_col=1, min_row=2, max_row=3)
bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(cats)
chart_ws.add_chart(bar_chart, "D2")  # Place at D2

# --- Passenger Class Distribution Data ---
pclass_counts = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    pclass_counts[row[2]] = pclass_counts.get(row[2], 0) + 1

chart_ws["A5"] = "Pclass"
chart_ws["B5"] = "Count"
for i, (cls, count) in enumerate(pclass_counts.items(), start=6):
    chart_ws[f"A{i}"] = cls
    chart_ws[f"B{i}"] = count

# Passenger Class Pie Chart
pie_chart = PieChart()
pie_chart.title = "Passenger Class Distribution"
data = Reference(chart_ws, min_col=2, min_row=5, max_row=5 + len(pclass_counts))
labels = Reference(chart_ws, min_col=1, min_row=6, max_row=5 + len(pclass_counts))
pie_chart.add_data(data, titles_from_data=True)
pie_chart.set_categories(labels)
chart_ws.add_chart(pie_chart, "D20")  # Place at D20 (space after first chart)

# --- Age Distribution Data ---
bins = [(0,10),(10,20),(20,30),(30,40),(40,50),(50,60),(60,70),(70,80)]
age_counts = {f"{low}-{high}": 0 for low, high in bins}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    age = row[4]
    if age is not None:
        for low, high in bins:
            if low <= age < high:
                age_counts[f"{low}-{high}"] += 1
                break

chart_ws["A15"] = "Age Range"
chart_ws["B15"] = "Count"
for i, (age_range, count) in enumerate(age_counts.items(), start=16):
    chart_ws[f"A{i}"] = age_range
    chart_ws[f"B{i}"] = count

# Age Distribution Bar Chart
age_chart = BarChart()
age_chart.title = "Age Distribution"
age_chart.x_axis.title = "Age Range"
age_chart.y_axis.title = "Count"
data = Reference(chart_ws, min_col=2, min_row=15, max_row=15 + len(age_counts))
cats = Reference(chart_ws, min_col=1, min_row=16, max_row=15 + len(age_counts))
age_chart.add_data(data, titles_from_data=True)
age_chart.set_categories(cats)
chart_ws.add_chart(age_chart, "D38")  # Place at D38 (space after pie chart)

# Save the workbook
wb.save("cleaned_titanic_report_with_charts.xlsx")


